import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# FILE CONSTANTS
EXCEL_FILE = "ToDo-List-App/List_Memory.xlsx"
TXT_FILE = "ToDo-List-App/Create_ID.txt"

# Create the Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    db = wb.active
    db.title = "ToDo List"
    db.append(["ID", "Task", "Status", "Reminder Date", "Reminder Time"])
    wb.save(EXCEL_FILE)

def gen_id_key():
    """Generate a unique ID for each task."""
    with open(TXT_FILE, 'r') as file:
        current_value = int(file.readline().strip().split('ID ')[-1])
        new_value = current_value + 1
        id_key = f"{current_value:06d}"

    with open(TXT_FILE, 'w') as file:
        file.write(f"ID {new_value:06d}\n")

    return id_key

def load_tasks():
    """Load tasks from the Excel file."""
    tasks_listbox.delete(0, tk.END)  # Clear current tasks
    wb = load_workbook(EXCEL_FILE)
    db = wb.active
    for row in db.iter_rows(min_row=2, values_only=True):
        task_id, task, status, reminder_date, reminder_time = row
        display_text = (
            f"{task} - {'Complete' if status == 'Complete' else 'Incomplete'}"
            f" (Reminder: {reminder_date} {reminder_time})"
        )
        tasks_listbox.insert(tk.END, display_text)
    wb.close()

def add_task():
    """Add a new task to the to-do list."""
    task = task_entry.get().strip()
    reminder_date = date_entry.get().strip()
    reminder_time = time_entry.get().strip()

    if not task:
        messagebox.showwarning("Warning", "Task cannot be empty.")
        return
    if not reminder_date or not reminder_time:
        messagebox.showwarning("Warning", "Reminder date and time cannot be empty.")
        return

    try:
        datetime.strptime(reminder_date, "%Y-%m-%d")  # Validate date format
        datetime.strptime(reminder_time, "%H:%M")  # Validate time format
    except ValueError:
        messagebox.showerror("Error", "Invalid date or time format.")
        return

    task_id = gen_id_key()

    wb = load_workbook(EXCEL_FILE)
    db = wb.active
    db.append([task_id, task, "Incomplete", reminder_date, reminder_time])
    wb.save(EXCEL_FILE)
    load_tasks()
    task_entry.delete(0, tk.END)
    date_entry.delete(0, tk.END)
    time_entry.delete(0, tk.END)

def get_selected_task_id():
    """Retrieve the ID of the selected task."""
    try:
        selected_index = tasks_listbox.curselection()[0]
        selected_task = tasks_listbox.get(selected_index)
        task_name = selected_task.split(" - ")[0]

        wb = load_workbook(EXCEL_FILE)
        db = wb.active
        for row in db.iter_rows(min_row=2, values_only=True):
            if row[1] == task_name:
                return row[0]
    except IndexError:
        messagebox.showwarning("Warning", "No task selected.")
    return None

def mark_complete():
    """Mark the selected task as complete."""
    task_id = get_selected_task_id()
    if not task_id:
        return

    wb = load_workbook(EXCEL_FILE)
    db = wb.active
    for row in db.iter_rows(min_row=2, values_only=False):
        if row[0].value == task_id:
            row[2].value = "Complete"
            break
    wb.save(EXCEL_FILE)
    load_tasks()

def mark_incomplete():
    """Mark the selected task as incomplete."""
    task_id = get_selected_task_id()
    if not task_id:
        return

    wb = load_workbook(EXCEL_FILE)
    db = wb.active
    for row in db.iter_rows(min_row=2, values_only=False):
        if row[0].value == task_id:
            row[2].value = "Incomplete"
            break
    wb.save(EXCEL_FILE)
    load_tasks()

def delete_task():
    """Delete the selected task."""
    task_id = get_selected_task_id()
    if not task_id:
        return

    wb = load_workbook(EXCEL_FILE)
    db = wb.active
    for row in db.iter_rows(min_row=2, values_only=False):
        if row[0].value == task_id:
            db.delete_rows(row[0].row)
            break
    wb.save(EXCEL_FILE)
    load_tasks()

# GUI Setup
root = tk.Tk()
root.title("To-Do List App with Reminders")

frame = tk.Frame(root)
frame.pack(pady=10)

task_label = tk.Label(frame, text="Task:")
task_label.grid(row=0, column=0, padx=5)

task_entry = tk.Entry(frame, width=30)
task_entry.grid(row=0, column=1, padx=5)

add_button = tk.Button(frame, text="Add Task", command=add_task)
add_button.grid(row=0, column=4, padx=5)

date_label = tk.Label(frame, text="Reminder Date (YYYY-MM-DD):")
date_label.grid(row=1, column=0, padx=5)

date_entry = tk.Entry(frame, width=15)
date_entry.grid(row=1, column=1, padx=5)

time_label = tk.Label(frame, text="Reminder Time (HH:MM):")
time_label.grid(row=1, column=2, padx=5)

time_entry = tk.Entry(frame, width=10)
time_entry.grid(row=1, column=3, padx=5)

tasks_listbox = tk.Listbox(root, width=70, height=15)
tasks_listbox.pack(pady=10)

buttons_frame = tk.Frame(root)
buttons_frame.pack(pady=10)

complete_button = tk.Button(buttons_frame, text="Mark Complete", command=mark_complete)
complete_button.grid(row=0, column=0, padx=5)

incomplete_button = tk.Button(buttons_frame, text="Mark Incomplete", command=mark_incomplete)
incomplete_button.grid(row=0, column=1, padx=5)

delete_button = tk.Button(buttons_frame, text="Delete Task", command=delete_task)
delete_button.grid(row=0, column=2, padx=5)

exit_button = tk.Button(buttons_frame, text="Exit", command=root.quit)
exit_button.grid(row=0, column=3, padx=5)

# Load tasks on startup
load_tasks()

root.mainloop()
